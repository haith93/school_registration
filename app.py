"""
School Registration Dashboard - لوحة تسجيل الطلاب
Streamlit app for processing and visualizing school registration form responses.
"""

import io
import pandas as pd
import streamlit as st
from openpyxl.styles import Border, Side, Alignment
from openpyxl.utils import get_column_letter
from streamlit_gsheets import GSheetsConnection

# ─────────────────────────────────────────────
#  COLUMN MAPPING
# ─────────────────────────────────────────────
COL = {
    "timestamp":           "Timestamp",
    "father_name":         "اسم والد التلميذ/ة (الاسم الثلاثي باللّغة العربيّة):",
    "father_name2":        "اسم الأب (الاسم الثلاثي):",
    "mother_name":         "اسم أم التلميذ/ة (الاسم الثلاثي باللّغة العربيّة):",
    "phone":               " رقم الهاتف المستخدم للتواصل مع المدرسة (بالأرقام الانكليزيّة):",
    "wants_registration":  "أريد تسجيل ولدي في ثانويّة رحاب الزّهراء(ع) للعام الدّراسيّ 2026-2027.",
    "child1_name":         "اسم الولد الأكبر (الاسم الثلاثي باللّغة العربيّة):",
    "child2_name":         "اسم الولد الثّاني (الاسم الثلاثي باللّغة العربيّة):",
    "child3_name":         "اسم الولد الثّالث (الاسم الثلاثي باللّغة العربيّة):",
    "child4_name":         "اسم الولد الرّابع (الاسم الثلاثي باللّغة العربيّة):",
    "child5_name":         "اسم الولد الخامس (الاسم الثلاثي باللّغة العربيّة):",
    "child1_grade":        "الصف الحالي للولد الأكبر:",
    "child2_grade":        ":الصف الحالي للولد الثّاني",
    "child3_grade":        "الصف الحالي للولد الثّالث:",
    "child4_grade":        "الصف الحالي للولد الرّابع:",
    "child5_grade":        "الصف الحالي للولد الخامس:",
    "no_reg_reason":       "في حال عدم الرّغبة في تسجيل ولدكم، يُرجى التّفضل بذكر السّبب.",
    "notes":               "هل لديكم أي ملاحظات اخرى؟",
}

# ─────────────────────────────────────────────
#  FULL GRADE LIST
# ─────────────────────────────────────────────
ALL_GRADES_ORDERED = [
    ("الروضة الأولى",              "الروضة الأولى",       False),
    ("الروضة الثانية",             "الروضة الثانية",      False),
    ("الروضة الثانية - مساند",     "الروضة الثانية",      True),
    ("الروضة الثالثة",             "الروضة الثالثة",      False),
    ("الروضة الثالثة - مساند",     "الروضة الثالثة",      True),
    ("الأول الأساسي",              "الأول الأساسي",       False),
    ("الأول الأساسي - مساند",      "الأول الأساسي",       True),
    ("الثاني الأساسي",             "الثاني الأساسي",      False),
    ("الثاني الأساسي - مساند",     "الثاني الأساسي",      True),
    ("الثالث الأساسي",             "الثالث الأساسي",      False),
    ("الثالث الأساسي - مساند",     "الثالث الأساسي",      True),
    ("الرابع الأساسي",             "الرابع الأساسي",      False),
    ("الرابع الأساسي - مساند",     "الرابع الأساسي",      True),
    ("الخامس الأساسي",             "الخامس الأساسي",      False),
    ("الخامس الأساسي - مساند",     "الخامس الأساسي",      True),
    ("السادس الأساسي",             "السادس الأساسي",      False),
    ("السادس الأساسي - مساند",     "السادس الأساسي",      True),
    ("السابع الأساسي",             "السابع الأساسي",      False),
    ("السابع الأساسي - مساند",     "السابع الأساسي",      True),
    ("الثامن الأساسي",             "الثامن الأساسي",      False),
    ("الثامن الأساسي - مساند",     "الثامن الأساسي",      True),
    ("التاسع الأساسي",             "التاسع الأساسي",      False),
    ("التاسع الأساسي - مساند",     "التاسع الأساسي",      True),
    ("العاشر",                     "العاشر",              False),
    ("الحادي عشر",                 "الحادي عشر",          False),
    ("الثاني عشر",                 "الثاني عشر",          False),
]

GRADE_DISPLAY_NAMES = [g[0] for g in ALL_GRADES_ORDERED]
GRADE_RANK = {g[0]: i for i, g in enumerate(ALL_GRADES_ORDERED)}

# ─────────────────────────────────────────────
#  CATEGORIES
# ─────────────────────────────────────────────
KG_GRADES      = {"الروضة الأولى", "الروضة الثانية", "الروضة الثالثة"}
MASNAD_GRADES  = {g[0] for g in ALL_GRADES_ORDERED if g[2]}
BASIC_GRADES   = {
    "الأول الأساسي", "الثاني الأساسي", "الثالث الأساسي",
    "الرابع الأساسي", "الخامس الأساسي", "السادس الأساسي",
}
SECONDARY_GRADES = {
    "السابع الأساسي", "الثامن الأساسي", "التاسع الأساسي",
    "العاشر", "الحادي عشر", "الثاني عشر",
}

def get_category(grade_display: str) -> str:
    if grade_display in MASNAD_GRADES:    return "تعليم مساند"
    if grade_display in KG_GRADES:        return "روضات"
    if grade_display in BASIC_GRADES:     return "اساسي (الأول – السادس)"
    if grade_display in SECONDARY_GRADES: return "متوسط وثانوي (السابع – الثاني عشر)"
    return "غير محدد"

CAT_ORDER = [
    "روضات",
    "تعليم مساند",
    "اساسي (الأول – السادس)",
    "متوسط وثانوي (السابع – الثاني عشر)",
    "غير محدد",
]

CAT_COLORS = {
    "روضات":                                "#fef9c3",
    "تعليم مساند":                          "#fee2e2",
    "اساسي (الأول – السادس)":              "#dcfce7",
    "متوسط وثانوي (السابع – الثاني عشر)":  "#dbeafe",
    "غير محدد":                             "#f3f4f6",
}

# ─────────────────────────────────────────────
#  CYCLES
# ─────────────────────────────────────────────
CYCLE_LABELS = {
    "الأول الأساسي":  "الحلقة الأولى (1–3)",
    "الثاني الأساسي": "الحلقة الأولى (1–3)",
    "الثالث الأساسي": "الحلقة الأولى (1–3)",
    "الرابع الأساسي": "الحلقة الثانية (4–6)",
    "الخامس الأساسي": "الحلقة الثانية (4–6)",
    "السادس الأساسي": "الحلقة الثانية (4–6)",
    "السابع الأساسي": "الحلقة الثالثة (7–9)",
    "الثامن الأساسي": "الحلقة الثالثة (7–9)",
    "التاسع الأساسي": "الحلقة الثالثة (7–9)",
    "العاشر":          "الحلقة الرابعة (10–12)",
    "الحادي عشر":      "الحلقة الرابعة (10–12)",
    "الثاني عشر":      "الحلقة الرابعة (10–12)",
    "الروضة الأولى":   "روضات",
    "الروضة الثانية":  "روضات",
    "الروضة الثالثة":  "روضات",
}

# ─────────────────────────────────────────────
#  HELPERS
# ─────────────────────────────────────────────
NULL_VALUES = {"لا", "لا يوجد", "none", "n/a", "-", "nan", "", "لا يوجد.", "مافي", "ما في", "/"}

def is_null_value(v):
    if pd.isna(v): return True
    return str(v).strip().lower() in NULL_VALUES

def normalize_grade(raw) -> str | None:
    if pd.isna(raw): return None
    import re
    s = re.sub(r'\s*-\s*', ' - ', str(raw).strip())
    if s in GRADE_RANK: return s
    base = re.sub(r'\s*-\s*مساند.*', '', s).strip()
    masnad_candidate = base + " - مساند"
    if masnad_candidate in GRADE_RANK: return masnad_candidate
    if base in GRADE_RANK: return base
    return s

def grade_base(grade_display: str) -> str:
    if grade_display and " - مساند" in grade_display:
        return grade_display.replace(" - مساند", "").strip()
    return grade_display

def grade_rank(g) -> int:
    return GRADE_RANK.get(g, 999)

def family_display_name(row):
    """Return a readable family identifier for display."""
    for k in ["father_name", "father_name2", "mother_name"]:
        v = row.get(COL[k])
        if not is_null_value(v):
            return str(v).strip()
    return "—"

# ─────────────────────────────────────────────
#  CORE DATA PROCESSING
# ─────────────────────────────────────────────
def load_and_process(data_input, is_live_sheet=False):
    try:
        if is_live_sheet:
            raw = data_input.copy() # It's already a pandas DataFrame
        else:
            raw = pd.read_excel(data_input) # It's an uploaded file
    except Exception as e:
        st.error(f"❌ خطأ في قراءة البيانات: {e}")
        return None, None, None, None, None
    # try:
    #     raw = pd.read_excel(uploaded_file)
    # except Exception as e:
    #     st.error(f"❌ خطأ في قراءة الملف: {e}")
    #     return None, None, None, None, None

    missing = [v for v in COL.values() if v not in raw.columns]
    if missing:
        st.error("❌ الأعمدة التالية غير موجودة في الملف:\n" + "\n".join(missing))
        return None, None, None, None, None

    df = raw.copy()
    total_submissions = len(df)

    # ── Family key ────────────────────────────────────────────────────────
    def get_family_key(row):
        for k in ["father_name", "father_name2", "mother_name"]:
            v = row.get(COL[k])
            if not is_null_value(v):
                return str(v).strip()
        return None

    df["_family_key"]    = df.apply(get_family_key, axis=1)
    df["_display_name"]  = df.apply(family_display_name, axis=1)
    df[COL["timestamp"]] = pd.to_datetime(df[COL["timestamp"]], errors="coerce")
    df = df.sort_values(COL["timestamp"], ascending=False)

    # ── Build duplicate report BEFORE dedup ───────────────────────────────
    # For every family_key that appears more than once, record all submissions
    dup_records = []
    grouped = df.groupby("_family_key", sort=False)
    for fkey, grp in grouped:
        if len(grp) < 2:
            continue
        grp_sorted = grp.sort_values(COL["timestamp"], ascending=False)
        kept_ts = grp_sorted.iloc[0][COL["timestamp"]]
        for i, (_, r) in enumerate(grp_sorted.iterrows()):
            dup_records.append({
                "اسم الأسرة":       fkey,
                "رقم الهاتف":       r.get(COL["phone"], ""),
                "وقت الإرسال":      r[COL["timestamp"]],
                "الحالة":           "✅ محتفظ به (الأحدث)" if i == 0 else "🗑️ محذوف (مكرر)",
                "يريد التسجيل":     r.get(COL["wants_registration"], ""),
                "ملاحظات":          r.get(COL["notes"], ""),
            })
    df_duplicates = pd.DataFrame(dup_records) if dup_records else pd.DataFrame(
        columns=["اسم الأسرة", "رقم الهاتف", "وقت الإرسال", "الحالة", "يريد التسجيل", "ملاحظات"]
    )

    # ── Deduplicate: keep latest per family ───────────────────────────────
    df_deduped = df.drop_duplicates(subset="_family_key", keep="first").copy()
    duplicates_removed = total_submissions - len(df_deduped)

    df_deduped["_wants_reg"] = (
        df_deduped[COL["wants_registration"]].astype(str).str.strip() == "نعم"
    )

    # ── Build per-family records ──────────────────────────────────────────
    child_slots = [
        ("child1_name", "child1_grade"),
        ("child2_name", "child2_grade"),
        ("child3_name", "child3_grade"),
        ("child4_name", "child4_grade"),
        ("child5_name", "child5_grade"),
    ]

    # Children missing their grade — collected for a separate report
    gradeless_records = []

    records = []
    for _, row in df_deduped.iterrows():
        children_valid    = []   # name + grade both present
        seen_names        = set()

        for name_key, grade_key in child_slots:
            cname  = row.get(COL[name_key])
            cgrade = row.get(COL[grade_key])

            if is_null_value(cname):
                continue                        # no name → skip slot entirely

            cname_clean = str(cname).strip()
            if cname_clean.lower() in seen_names:
                continue
            seen_names.add(cname_clean.lower())

            cgrade_norm = normalize_grade(cgrade)

            # ── NEW: grade missing → log to gradeless report, skip from counts ──
            if is_null_value(cgrade) or cgrade_norm is None:
                gradeless_records.append({
                    "اسم الأسرة":    row["_family_key"] or "—",
                    "رقم الهاتف":    row.get(COL["phone"], ""),
                    "اسم الطالب":    cname_clean,
                    "الصف المُدخل":  str(cgrade).strip() if not pd.isna(cgrade) else "—",
                    "يريد التسجيل":  "نعم" if row["_wants_reg"] else "لا",
                })
                continue                        # excluded from all statistics

            children_valid.append((cname_clean, cgrade_norm))

        # Oldest child = highest grade rank among VALID children
        children_sorted = sorted(children_valid, key=lambda x: grade_rank(x[1]), reverse=True)
        oldest = children_sorted[0] if children_sorted else (None, None)

        oldest_base = grade_base(oldest[1]) if oldest[1] else None
        cycle = CYCLE_LABELS.get(oldest_base, "روضات")

        records.append({
            "اسم الأب":         row.get(COL["father_name"], row.get(COL["father_name2"])),
            "اسم الأم":         row.get(COL["mother_name"]),
            "رقم الهاتف":       row.get(COL["phone"]),
            "يريد التسجيل":     "نعم" if row["_wants_reg"] else "لا",
            "_wants_reg":        row["_wants_reg"],
            "الابن الأكبر":     oldest[0] or "—",
            "صف الابن الأكبر":  oldest[1] or "—",
            "الحلقة":            cycle,
            "عدد الأبناء":       len(children_valid),
            # "أسماء الأبناء":    "، ".join([n for n, _ in children_valid]) or "—",
            "أسماء الأبناء":    "، ".join([f"{n} ({g})" for n, g in children_valid]) or "—",
            "ملاحظات":           row.get(COL["notes"], ""),
            "سبب عدم التسجيل":  row.get(COL["no_reg_reason"], ""),
            "_children":         children_valid,
            "_family_key":       row["_family_key"],
        })

    df_families  = pd.DataFrame(records)
    df_gradeless = pd.DataFrame(gradeless_records) if gradeless_records else pd.DataFrame(
        columns=["اسم الأسرة", "رقم الهاتف", "اسم الطالب", "الصف المُدخل", "يريد التسجيل"]
    )

    # ── Statistics ────────────────────────────────────────────────────────
    total_unique = len(df_families)
    wants_yes    = int(df_families["_wants_reg"].sum())
    wants_no     = total_unique - wants_yes

    grade_counter = {}
    for _, row in df_families.iterrows():
        for name, grade in row["_children"]:
            g = grade or "غير محدد"
            grade_counter[g] = grade_counter.get(g, 0) + 1

    cat_counter = {}
    for g, cnt in grade_counter.items():
        cat = get_category(g)
        cat_counter[cat] = cat_counter.get(cat, 0) + cnt

    cycle_counter = (
        df_families[df_families["_wants_reg"]]["الحلقة"]
        .value_counts().to_dict()
    )

    total_students_wanting_reg = sum(
        row["عدد الأبناء"]
        for _, row in df_families[df_families["_wants_reg"]].iterrows()
    )

    stats = {
        "total_submissions":          total_submissions,
        "duplicates_removed":         duplicates_removed,
        "total_unique_families":      total_unique,
        "families_want_yes":          wants_yes,
        "families_want_no":           wants_no,
        "pct_yes":                    (wants_yes / total_unique * 100) if total_unique else 0,
        "pct_no":                     (wants_no  / total_unique * 100) if total_unique else 0,
        "total_children":             sum(grade_counter.values()),
        "total_students_wanting_reg": total_students_wanting_reg,
        "gradeless_count":            len(df_gradeless),
        "grade_breakdown":            grade_counter,
        "cat_breakdown":              cat_counter,
        "cycle_breakdown":            cycle_counter,
    }

    return raw, df_families, stats, df_duplicates, df_gradeless


# ─────────────────────────────────────────────
#  EXPORT HELPERS
# ─────────────────────────────────────────────
# ─────────────────────────────────────────────
#  EXPORT HELPERS
# ─────────────────────────────────────────────
def apply_excel_formatting(worksheet, df):
    """Applies borders, auto-width, and RTL layout to a worksheet."""
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    align_center_right = Alignment(horizontal='right', vertical='center', wrap_text=True)

    # 1. Calculate and set column widths
    for i, col in enumerate(df.columns):
        # Find the max length of the data in the column, or 0 if empty
        # max_data_len = df[col].astype(str).map(len).max() if not df.empty else 0
        max_data_len = df[col].astype(str).str.len().max() if not df.empty else 0
        header_len = len(str(col))
        
        # Determine the final width (+4 for some breathing room)
        # Cap at 60 so super long notes don't make the column insanely wide
        final_width = min(max(max_data_len, header_len) + 4, 60) 
        
        col_letter = get_column_letter(i + 1)
        worksheet.column_dimensions[col_letter].width = final_width

    # 2. Apply borders and alignment to every cell
    for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row, min_col=1, max_col=worksheet.max_column):
        for cell in row:
            cell.border = thin_border
            cell.alignment = align_center_right

    # 3. Make the sheet open natively as Right-to-Left
    worksheet.sheet_view.rightToLeft = True

def to_excel_bytes(df: pd.DataFrame) -> bytes:
    buf = io.BytesIO()
    export_cols = [c for c in df.columns if not str(c).startswith("_")]
    df_clean = df[export_cols]
    
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        df_clean.to_excel(writer, index=False, sheet_name="البيانات")
        apply_excel_formatting(writer.sheets["البيانات"], df_clean)
        
    return buf.getvalue()


def stats_to_excel_bytes(stats: dict) -> bytes:
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        
        # Sheet 1: Summary
        df_summary = pd.DataFrame({
            "البند": [
                "إجمالي النماذج المستلمة", "النماذج المكررة المحذوفة", "الأسر الفريدة",
                "الأسر الراغبة في التسجيل", "الأسر الغير راغبة في التسجيل",
                "نسبة الراغبين (%)", "نسبة الغير راغبين (%)",
                "إجمالي الأطفال (بصف محدد)", "الطلاب الراغبون في التسجيل", "طلاب بدون صف (مستثنون)",
            ],
            "القيمة": [
                stats["total_submissions"], stats["duplicates_removed"], stats["total_unique_families"],
                stats["families_want_yes"], stats["families_want_no"],
                round(stats["pct_yes"], 1), round(stats["pct_no"], 1),
                stats["total_children"], stats["total_students_wanting_reg"], stats["gradeless_count"],
            ],
        })
        df_summary.to_excel(writer, sheet_name="ملخص", index=False)
        apply_excel_formatting(writer.sheets["ملخص"], df_summary)

        # Sheet 2: Grades
        df_grades = pd.DataFrame(list(stats["grade_breakdown"].items()), columns=["الصف", "عدد الأطفال"])
        df_grades["القسم / Section"] = df_grades["الصف"].map(get_category)
        df_grades["_ord"] = df_grades["الصف"].map(lambda g: GRADE_RANK.get(g, 999))
        df_grades = df_grades.sort_values("_ord").drop(columns="_ord")
        df_grades.to_excel(writer, sheet_name="توزيع الصفوف", index=False)
        apply_excel_formatting(writer.sheets["توزيع الصفوف"], df_grades)

        # Sheet 3: Categories
        df_cats = pd.DataFrame(list(stats["cat_breakdown"].items()), columns=["الفئة", "عدد الأطفال"])
        df_cats.to_excel(writer, sheet_name="توزيع الفئات", index=False)
        apply_excel_formatting(writer.sheets["توزيع الفئات"], df_cats)

        # Sheet 4: Cycles
        df_cycles = pd.DataFrame(list(stats["cycle_breakdown"].items()), columns=["الحلقة", "عدد الأسر"])
        df_cycles.to_excel(writer, sheet_name="توزيع الحلقات", index=False)
        apply_excel_formatting(writer.sheets["توزيع الحلقات"], df_cycles)

    return buf.getvalue()
# def to_excel_bytes(df: pd.DataFrame) -> bytes:
#     buf = io.BytesIO()
#     export_cols = [c for c in df.columns if not str(c).startswith("_")]
#     df[export_cols].to_excel(buf, index=False)
#     return buf.getvalue()


# def stats_to_excel_bytes(stats: dict) -> bytes:
#     buf = io.BytesIO()
#     with pd.ExcelWriter(buf, engine="openpyxl") as writer:
#         pd.DataFrame({
#             "البند": [
#                 "إجمالي النماذج المستلمة",
#                 "النماذج المكررة المحذوفة",
#                 "الأسر الفريدة",
#                 "الأسر الراغبة في التسجيل",
#                 "الأسر الغير راغبة في التسجيل",
#                 "نسبة الراغبين (%)",
#                 "نسبة الغير راغبين (%)",
#                 "إجمالي الأطفال (بصف محدد)",
#                 "الطلاب الراغبون في التسجيل",
#                 "طلاب بدون صف (مستثنون)",
#             ],
#             "القيمة": [
#                 stats["total_submissions"],
#                 stats["duplicates_removed"],
#                 stats["total_unique_families"],
#                 stats["families_want_yes"],
#                 stats["families_want_no"],
#                 round(stats["pct_yes"], 1),
#                 round(stats["pct_no"], 1),
#                 stats["total_children"],
#                 stats["total_students_wanting_reg"],
#                 stats["gradeless_count"],
#             ],
#         }).to_excel(writer, sheet_name="ملخص", index=False)

#         grade_df = pd.DataFrame(
#             list(stats["grade_breakdown"].items()), columns=["الصف", "عدد الأطفال"]
#         )
#         grade_df["القسم / Section"] = grade_df["الصف"].map(get_category)
#         grade_df["_ord"] = grade_df["الصف"].map(lambda g: GRADE_RANK.get(g, 999))
#         grade_df.sort_values("_ord").drop(columns="_ord").to_excel(
#             writer, sheet_name="توزيع الصفوف", index=False
#         )

#         pd.DataFrame(
#             list(stats["cat_breakdown"].items()), columns=["الفئة", "عدد الأطفال"]
#         ).to_excel(writer, sheet_name="توزيع الفئات", index=False)

#         pd.DataFrame(
#             list(stats["cycle_breakdown"].items()), columns=["الحلقة", "عدد الأسر"]
#         ).to_excel(writer, sheet_name="توزيع الحلقات", index=False)

#     return buf.getvalue()


# ─────────────────────────────────────────────
#  STREAMLIT UI
# ─────────────────────────────────────────────
st.set_page_config(
    page_title="لوحة تسجيل الطلاب",
    page_icon="🏫",
    layout="wide",
    initial_sidebar_state="expanded",
)

st.markdown("""
<style>
body, .stApp { direction: rtl; }
.stDataFrame { direction: rtl; }
            
section[data-testid="stSidebar"][aria-expanded="false"] {
    min-width: 0px !important;
    width: 0px !important;
    opacity: 0 !important;
    border: none !important;
    box-shadow: none !important;
    transform: translateX(100%) !important; /* Gently moves it off-screen without altering margins */
}

[data-testid="collapsedControl"] {
    z-index: 99999 !important;
}
            
.metric-card {
    background: #eef2ff;
    border-radius: 12px;
    padding: 16px 18px;
    text-align: center;
    border: 1px solid #c7d2fe;
    margin-bottom: 8px;
}
.metric-card-green {
    background: #dcfce7;
    border-radius: 12px;
    padding: 16px 18px;
    text-align: center;
    border: 1px solid #86efac;
    margin-bottom: 8px;
}
.metric-card-warn {
    background: #fff7ed;
    border-radius: 12px;
    padding: 16px 18px;
    text-align: center;
    border: 1px solid #fdba74;
    margin-bottom: 8px;
}
.metric-label { font-size: 0.82rem; color: #555; margin-bottom: 4px; }
.metric-value { font-size: 1.9rem; font-weight: 700; color: #1e3a8a; }
.metric-sub   { font-size: 0.78rem; color: #888; }
.cat-card {
    border-radius: 10px;
    padding: 14px 16px;
    margin-bottom: 8px;
    border: 1px solid #e2e8f0;
}
h1, h2, h3 { color: #1e3a8a; }
.dup-kept   { background: #dcfce7; padding: 2px 8px; border-radius: 6px; font-size:0.82rem; }
.dup-del    { background: #fee2e2; padding: 2px 8px; border-radius: 6px; font-size:0.82rem; }
</style>
""", unsafe_allow_html=True)

st.title("🏫 لوحة تسجيل الطلاب – ثانويّة رحاب الزّهراء(ع)")
st.markdown("---")

with st.sidebar:
    st.header("⚙️ مصدر البيانات")
    
    # Give yourself a toggle
    data_source = st.radio(
        "اختر طريقة جلب البيانات:",
        ["جلب مباشر من Google Sheets 🌐", "رفع ملف Excel 📂"]
    )
    
    st.markdown("---")
    
    uploaded_file = None
    live_df = None
    
    if "Google Sheets" in data_source:
        sheet_url = st.text_input("رابط Google Sheet", placeholder="الصق الرابط هنا...")
        if sheet_url:
            try:
                # Connect to Google Sheets and read the data
                conn = st.connection("gsheets", type=GSheetsConnection)
                # ttl=60 means it caches the data for 60 seconds so it doesn't overload the API
                live_df = conn.read(spreadsheet=sheet_url, ttl=60)
                st.success("✅ تم جلب البيانات المباشرة بنجاح!")
            except Exception as e:
                st.error(f"فشل في الاتصال بـ Google Sheets: {e}")
                st.stop()
        else:
            st.info("👈 الرجاء إدخال رابط Google Sheet الخاص بالاستجابات.")
            st.stop()
            
    else:
        uploaded_file = st.file_uploader("ارفع ملف Excel (نموذج الاستجابات)", type=["xlsx", "xls"])
        if uploaded_file is None:
            st.info("👈 الرجاء رفع ملف Excel للبدء.")
            st.stop()

# Decide what to pass into your processing function
if "Google Sheets" in data_source and live_df is not None:
    raw_df, df_families, stats, df_duplicates, df_gradeless = load_and_process(live_df, is_live_sheet=True)
elif uploaded_file is not None:
    raw_df, df_families, stats, df_duplicates, df_gradeless = load_and_process(uploaded_file, is_live_sheet=False)
else:
    st.stop()

# with st.sidebar:
#     st.header("📂 رفع الملف")
#     uploaded = st.file_uploader("ارفع ملف Excel (نموذج الاستجابات)", type=["xlsx", "xls"])
#     st.markdown("---")
#     st.info("💡 ارفع نسخة جديدة في أي وقت وسيتم إعادة الحساب تلقائياً.")

# if uploaded is None:
#     st.info("👈 الرجاء رفع ملف Excel من الشريط الجانبي للبدء.")
#     st.stop()

# raw_df, df_families, stats, df_duplicates, df_gradeless = load_and_process(uploaded)
# if df_families is None:
#     st.stop()

# # ── Alert banners ──────────────────────────────────────────────────────────
# if stats["duplicates_removed"] > 0:
#     st.warning(
#         f"⚠️ تم اكتشاف **{stats['duplicates_removed']}** نموذج مكرر وحُذف. "
#         f"راجع تبويب **تقرير المكررات** للتفاصيل."
#     )
# if stats["gradeless_count"] > 0:
#     st.warning(
#         f"⚠️ تم استثناء **{stats['gradeless_count']}** طالب لعدم تحديد صفهم. "
#         f"راجع تبويب **طلاب بدون صف** للتفاصيل."
#     )

# ══════════════════════════════════════════════════════════════
tab1, tab2, tab3, tab4, tab5, tab6 = st.tabs([
    "📊 الملخص العام",
    "👨‍👩‍👧 جدول الأسر",
    "📋 لوحة المشرف",
    "🔁 تقرير المكررات",
    "⚠️ طلاب بدون صف",
    "📤 تصدير",
])

# ══════════════════════════════════════════════════════════════
#  TAB 1 – SUMMARY
# ══════════════════════════════════════════════════════════════
with tab1:
    st.subheader("إحصاءات عامة")

    c1, c2, c3 = st.columns(3)
    c4, c5, c6 = st.columns(3)

    def metric(col, label, value, sub="", style=""):
        css = f"metric-card{'-' + style if style else ''}"
        col.markdown(f"""
        <div class="{css}">
            <div class="metric-label">{label}</div>
            <div class="metric-value">{value}</div>
            <div class="metric-sub">{sub}</div>
        </div>""", unsafe_allow_html=True)

    metric(c1, "إجمالي النماذج",        stats["total_submissions"])
    metric(c2, "مكررة محذوفة",           stats["duplicates_removed"])
    metric(c3, "أسر فريدة",             stats["total_unique_families"])
    metric(c4, "يريدون التسجيل",        stats["families_want_yes"],          f'{stats["pct_yes"]:.1f}%')
    metric(c5, "لا يريدون التسجيل",      stats["families_want_no"],           f'{stats["pct_no"]:.1f}%')
    metric(c6, "إجمالي الطلاب الراغبين", stats["total_students_wanting_reg"], "طالب / طالبة", "green")

    # Gradeless warning card inline
    if stats["gradeless_count"] > 0:
        st.markdown(f"""
        <div class="metric-card-warn" style="margin-top:8px;">
            <div class="metric-label">طلاب مستثنون (بدون صف)</div>
            <div class="metric-value">{stats["gradeless_count"]}</div>
            <div class="metric-sub">غير محسوبين في أي إحصاء</div>
        </div>""", unsafe_allow_html=True)

    st.markdown("<br>", unsafe_allow_html=True)

    st.subheader("📂 توزيع الأطفال حسب الفئة")
    cat_cols = st.columns(len(CAT_ORDER))
    total_ch = stats["total_children"] or 1
    for i, cat in enumerate(CAT_ORDER):
        cnt = stats["cat_breakdown"].get(cat, 0)
        pct = cnt / total_ch * 100
        bg  = CAT_COLORS.get(cat, "#f3f4f6")
        cat_cols[i].markdown(f"""
        <div class="cat-card" style="background:{bg};">
            <div class="metric-label">{cat}</div>
            <div class="metric-value">{cnt}</div>
            <div class="metric-sub">{pct:.1f}% من الأطفال</div>
        </div>""", unsafe_allow_html=True)

    st.markdown("<br>", unsafe_allow_html=True)

    col_a, col_b = st.columns(2)
    with col_a:
        st.subheader("📚 توزيع الأطفال حسب الصف")
        grade_df = pd.DataFrame(
            list(stats["grade_breakdown"].items()), columns=["الصف", "عدد الأطفال"]
        )
        grade_df["القسم / Section"] = grade_df["الصف"].map(get_category)
        grade_df["_ord"] = grade_df["الصف"].map(lambda g: GRADE_RANK.get(g, 999))
        grade_df = grade_df.sort_values("_ord").drop(columns="_ord")
        st.dataframe(grade_df, use_container_width=True, hide_index=True)

    with col_b:
        st.subheader("🔁 توزيع الأسر حسب الحلقة")
        cycle_df = pd.DataFrame(
            list(stats["cycle_breakdown"].items()), columns=["الحلقة", "عدد الأسر"]
        ).sort_values("عدد الأسر", ascending=False)
        st.dataframe(cycle_df, use_container_width=True, hide_index=True)

    st.metric("إجمالي الأطفال في الملف (بصف محدد)", stats["total_children"])


# ══════════════════════════════════════════════════════════════
#  TAB 2 – FAMILIES TABLE
# ══════════════════════════════════════════════════════════════
with tab2:
    st.subheader("جدول الأسر مع الفلاتر")

    f1, f2, f3, f4, f5 = st.columns(5)
    with f1:
        reg_filter   = st.selectbox("التسجيل", ["الكل", "نعم", "لا"])
    with f2:
        all_cycles   = ["الكل"] + sorted(df_families["الحلقة"].unique().tolist())
        cycle_filter = st.selectbox("الحلقة", all_cycles)
    with f3:
        cat_options  = ["الكل"] + CAT_ORDER[:-1]
        cat_filter   = st.selectbox("الفئة", cat_options)
    with f4:
        all_g_opts   = ["الكل"] + sorted(stats["grade_breakdown"].keys(),
                                          key=lambda g: GRADE_RANK.get(g, 999))
        grade_filter = st.selectbox("الصف", all_g_opts)
    with f5:
        search_fam   = st.text_input("🔍 بحث اسم الأسرة")

    filtered = df_families.copy()
    if reg_filter  != "الكل": filtered = filtered[filtered["يريد التسجيل"] == reg_filter]
    if cycle_filter != "الكل": filtered = filtered[filtered["الحلقة"] == cycle_filter]
    if cat_filter  != "الكل":
        filtered = filtered[filtered["_children"].apply(
            lambda ch: any(get_category(g) == cat_filter for _, g in ch))]
    if grade_filter != "الكل":
        filtered = filtered[filtered["_children"].apply(
            lambda ch: any(g == grade_filter for _, g in ch))]
    if search_fam:
        mask = (
            filtered["اسم الأب"].astype(str).str.contains(search_fam, na=False) |
            filtered["اسم الأم"].astype(str).str.contains(search_fam, na=False)
        )
        filtered = filtered[mask]

    st.caption(f"عرض {len(filtered)} أسرة من أصل {len(df_families)}")
    export_cols = [c for c in filtered.columns if not c.startswith("_")]
    st.dataframe(filtered[export_cols], use_container_width=True, hide_index=True)
    st.download_button(
        "⬇️ تصدير الجدول الحالي (Excel)",
        data=to_excel_bytes(filtered),
        file_name="الأسر_المفلترة.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


# ══════════════════════════════════════════════════════════════
#  TAB 3 – SUPERVISOR
# ══════════════════════════════════════════════════════════════
with tab3:
    st.subheader("لوحة المشرف – اختر الحلقة")

    available_cycles = sorted(df_families["الحلقة"].unique().tolist())
    selected_cycle   = st.selectbox("اختر الحلقة", ["-- اختر --"] + available_cycles)

    if selected_cycle == "-- اختر --":
        st.info("الرجاء اختيار حلقة من القائمة أعلاه.")
    else:
        cycle_data = df_families[df_families["الحلقة"] == selected_cycle]
        cycle_reg  = cycle_data[cycle_data["_wants_reg"]]
        cycle_no   = cycle_data[~cycle_data["_wants_reg"]]

        ca, cb = st.columns(2)
        ca.metric("الأسر الراغبة في التسجيل", len(cycle_reg))
        cb.metric("الأسر الغير راغبة",         len(cycle_no))

        sup_cols = [
            "اسم الأب", "اسم الأم", "رقم الهاتف",
            "يريد التسجيل", "الابن الأكبر", "صف الابن الأكبر",
            "الحلقة", "عدد الأبناء", "أسماء الأبناء", "ملاحظات",
        ]

        if len(cycle_reg):
            st.markdown("#### ✅ الأسر الراغبة في التسجيل")
            st.dataframe(cycle_reg[sup_cols], use_container_width=True, hide_index=True)

        if len(cycle_no):
            with st.expander(f"❌ الأسر الغير راغبة ({len(cycle_no)})"):
                no_cols = ["اسم الأب", "اسم الأم", "رقم الهاتف",
                           "يريد التسجيل", "سبب عدم التسجيل", "ملاحظات"]
                st.dataframe(cycle_no[no_cols], use_container_width=True, hide_index=True)

        st.download_button(
            f"⬇️ تصدير بيانات {selected_cycle}",
            data=to_excel_bytes(cycle_data),
            file_name=f"بيانات_{selected_cycle}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )


# ══════════════════════════════════════════════════════════════
#  TAB 4 – DUPLICATE REPORT
# ══════════════════════════════════════════════════════════════
with tab4:
    st.subheader("🔁 تقرير النماذج المكررة")

    if df_duplicates.empty:
        st.success("✅ لا توجد نماذج مكررة في الملف.")
    else:
        # Summary numbers
        n_families_duped = df_duplicates["اسم الأسرة"].nunique()
        n_removed        = stats["duplicates_removed"]

        d1, d2 = st.columns(2)
        d1.metric("الأسر التي أرسلت أكثر من مرة", n_families_duped)
        d2.metric("النماذج المحذوفة إجمالاً",       n_removed)

        st.markdown("---")
        st.markdown(
            "الجدول أدناه يعرض **جميع النماذج** لكل أسرة أرسلت أكثر من مرة. "
            "النموذج المحتفظ به هو **الأحدث تاريخاً**، والباقي يُعدّ مكرراً محذوفاً."
        )

        # Group-by view: one expander per family
        for fkey in df_duplicates["اسم الأسرة"].unique():
            fdata = df_duplicates[df_duplicates["اسم الأسرة"] == fkey].sort_values(
                "وقت الإرسال", ascending=False
            )
            kept_label  = fdata.iloc[0]["الحالة"]
            n_total     = len(fdata)
            with st.expander(f"👨‍👩‍👧 {fkey}  —  {n_total} نماذج"):
                st.dataframe(
                    fdata.drop(columns="اسم الأسرة"),
                    use_container_width=True,
                    hide_index=True,
                )

        st.markdown("---")
        st.download_button(
            "⬇️ تحميل تقرير المكررات (Excel)",
            data=to_excel_bytes(df_duplicates),
            file_name="تقرير_المكررات.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )


# ══════════════════════════════════════════════════════════════
#  TAB 5 – GRADELESS STUDENTS
# ══════════════════════════════════════════════════════════════
with tab5:
    st.subheader("⚠️ طلاب مُدخَلة أسماؤهم بدون صف دراسي")

    if df_gradeless.empty:
        st.success("✅ جميع الطلاب لديهم صف دراسي محدد.")
    else:
        st.info(
            f"**{len(df_gradeless)} طالب/ة** تم تسجيل اسمهم في النموذج لكن دون تحديد الصف. "
            "هؤلاء **مستثنون تماماً** من جميع الإحصاءات والجداول. "
            "إخوتهم الذين لديهم صف محدد **لا يزالون محسوبين** بشكل طبيعي."
        )
        st.markdown("---")

        # Filter by family
        gl_families = ["الكل"] + sorted(df_gradeless["اسم الأسرة"].unique().tolist())
        gl_fam_sel  = st.selectbox("فلترة حسب الأسرة", gl_families, key="gl_fam")

        gl_display = df_gradeless if gl_fam_sel == "الكل" else df_gradeless[
            df_gradeless["اسم الأسرة"] == gl_fam_sel
        ]
        st.caption(f"عرض {len(gl_display)} طالب")
        st.dataframe(gl_display, use_container_width=True, hide_index=True)

        st.download_button(
            "⬇️ تحميل تقرير الطلاب بدون صف (Excel)",
            data=to_excel_bytes(df_gradeless),
            file_name="طلاب_بدون_صف.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )


# ══════════════════════════════════════════════════════════════
#  TAB 6 – EXPORTS
# ══════════════════════════════════════════════════════════════
with tab6:
    st.subheader("تصدير البيانات")

    c1, c2, c3 = st.columns(3)
    with c1:
        st.markdown("#### 📋 جميع الأسر")
        st.download_button(
            "⬇️ تحميل جميع البيانات",
            data=to_excel_bytes(df_families),
            file_name="جميع_الأسر.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )
    with c2:
        st.markdown("#### 📊 ملخص الإحصاءات")
        st.download_button(
            "⬇️ تحميل الإحصاءات",
            data=stats_to_excel_bytes(stats),
            file_name="ملخص_الإحصاءات.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )
    with c3:
        st.markdown("#### 🔢 البيانات الخام")
        raw_buf = io.BytesIO()
        raw_df.to_excel(raw_buf, index=False)
        st.download_button(
            "⬇️ تحميل البيانات الخام",
            data=raw_buf.getvalue(),
            file_name="البيانات_الخام.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )

    st.markdown("---")
    st.markdown("#### 🔁 تقارير المعالجة")
    ecol1, ecol2 = st.columns(2)
    with ecol1:
        st.download_button(
            f"⬇️ تقرير المكررات ({stats['duplicates_removed']} نموذج)",
            data=to_excel_bytes(df_duplicates),
            file_name="تقرير_المكررات.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )
    with ecol2:
        st.download_button(
            f"⬇️ طلاب بدون صف ({stats['gradeless_count']} طالب)",
            data=to_excel_bytes(df_gradeless),
            file_name="طلاب_بدون_صف.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )

    st.markdown("---")
    st.markdown("#### 📁 تصدير كل حلقة على حدة")
    for cycle in sorted(df_families["الحلقة"].unique()):
        cdf = df_families[df_families["الحلقة"] == cycle]
        st.download_button(
            f"⬇️ {cycle}  ({len(cdf)} أسرة)",
            data=to_excel_bytes(cdf),
            file_name=f"{cycle}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key=f"exp_{cycle}",
        )

    st.markdown("---")
    st.markdown("#### 📁 تصدير كل فئة على حدة")
    for cat in CAT_ORDER:
        cat_fams = df_families[df_families["_children"].apply(
            lambda ch: any(get_category(g) == cat for _, g in ch)
        )]
        if len(cat_fams):
            st.download_button(
                f"⬇️ {cat}  ({len(cat_fams)} أسرة)",
                data=to_excel_bytes(cat_fams),
                file_name=f"فئة_{cat}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key=f"exp_cat_{cat}",
            )